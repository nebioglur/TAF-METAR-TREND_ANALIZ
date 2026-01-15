# -*- coding: utf-8 -*-
import tkinter as tk
from tkinter import ttk, messagebox, filedialog
import tkinter.font as tkfont
from datetime import datetime, timedelta, timezone
import threading
import pandas as pd
import re
import math
import requests
from bs4 import BeautifulSoup
from tkcalendar import DateEntry
import urllib3
import time
import json
import os
import RASATLAR
import TAF_METAR_TREND
from ayarlar import STATION, WMO_ID, TURKEY_STATIONS, TURKEY_BORDER
from veri_isleme import process_data

# SSL Hatalarını Gizle
urllib3.disable_warnings(urllib3.exceptions.InsecureRequestWarning)

robot = TAF_METAR_TREND.HavacilikRobotModulu()

# ================== GUI ==================
class App(tk.Tk):
    def __init__(self):
        super().__init__()
        self.config_file = "user_settings.json"
        self.saved_station = STATION
        self.saved_wmo = WMO_ID
        self.load_config()
        
        self.title(f"OGIMET ICAO ANALİZ - {self.saved_station}")
        self.geometry("1200x700")
        self.state('zoomed')
        self.configure(bg="#2b2b2b")
        self.full_df = None
        self.auto_refresh_var = tk.BooleanVar()
        self.bg_scan_var = tk.BooleanVar()
        self.incompatible_df = pd.DataFrame() # Uyumsuzluk takip verisi
        self.monitor_window = None
        self.monitor_tree = None
        self.refresh_job = None
        self.tree_tooltips = {}
        self.tooltip_window = None
        self.last_tooltip_item = None
        self.auto_save_monitor_loop() # Otomatik kayıt döngüsünü başlat
        self.setup_ui()
        
    def load_config(self):
        if os.path.exists(self.config_file):
            try:
                with open(self.config_file, "r", encoding="utf-8") as f:
                    data = json.load(f)
                    self.saved_station = data.get("station", STATION)
                    self.saved_wmo = data.get("wmo", WMO_ID)
            except: pass

    def save_config(self):
        data = {
            "station": self.ent_station.get().strip().upper(),
            "wmo": self.ent_wmo.get().strip()
        }
        try:
            with open(self.config_file, "w", encoding="utf-8") as f:
                json.dump(data, f)
        except: pass

    def setup_ui(self):
        style = ttk.Style()
        style.theme_use('clam')
        style.configure("Treeview", background="#263238", foreground="#eceff1", fieldbackground="#263238", rowheight=45, font=("Segoe UI", 10))
        style.configure("Treeview.Heading", background="#37474f", foreground="white", font=("Segoe UI", 11, "bold"))
        style.map("Treeview", background=[('selected', '#00bcd4')])

        top_frame = tk.Frame(self, bg="#2b2b2b")
        top_frame.pack(fill="x", padx=10, pady=10)
        
        tk.Label(top_frame, text="ICAO:", bg="#2b2b2b", fg="#aaaaaa").pack(side="left")
        self.ent_station = tk.Entry(top_frame, width=5, bg="#1e1e1e", fg="white", insertbackground="white")
        self.ent_station.insert(0, self.saved_station)
        self.ent_station.pack(side="left", padx=(2, 10))

        tk.Label(top_frame, text="WMO:", bg="#2b2b2b", fg="#aaaaaa").pack(side="left")
        self.ent_wmo = tk.Entry(top_frame, width=6, bg="#1e1e1e", fg="white", insertbackground="white")
        self.ent_wmo.insert(0, self.saved_wmo)
        self.ent_wmo.pack(side="left", padx=(2, 10))
        
        now = datetime.now()
        self.ent_start = DateEntry(top_frame, width=10, background='#0078D7', foreground='white', borderwidth=2, date_pattern='dd.mm.yyyy')
        self.ent_start.set_date(now - timedelta(days=1))
        self.ent_start.pack(side="left", padx=5)
        
        # Başlangıç Saati
        self.cb_start_hour = ttk.Combobox(top_frame, values=[f"{i:02d}" for i in range(24)], width=3, state="readonly")
        self.cb_start_hour.set("00")
        self.cb_start_hour.pack(side="left", padx=2)
        
        self.ent_end = DateEntry(top_frame, width=10, background='#0078D7', foreground='white', borderwidth=2, date_pattern='dd.mm.yyyy')
        self.ent_end.set_date(now)
        self.ent_end.pack(side="left", padx=5)
        
        # Bitiş Saati
        self.cb_end_hour = ttk.Combobox(top_frame, values=[f"{i:02d}" for i in range(24)], width=3, state="readonly")
        self.cb_end_hour.set("23")
        self.cb_end_hour.pack(side="left", padx=2)
        
        tk.Button(top_frame, text="VERİ ÇEK & ANALİZ ET", command=self.start_process, 
                  bg="#0078D7", fg="white", font=("Segoe UI", 10, "bold"), relief="flat").pack(side="left", padx=10)
        
        tk.Button(top_frame, text="EXCEL RAPOR", command=self.export_to_excel, 
                  bg="#43A047", fg="white", font=("Segoe UI", 10, "bold"), relief="flat").pack(side="left", padx=5)
        
        # TÜRKİYE HARİTASI BUTONU
        tk.Button(top_frame, text="TÜRKİYE HARİTASI", command=self.open_turkey_map, 
                  bg="#D81B60", fg="white", font=("Segoe UI", 10, "bold"), relief="flat").pack(side="left", padx=5)
        
        # Arka Plan Tarama Checkbox
        self.chk_bg_scan = tk.Checkbutton(top_frame, text="TR Arka Plan Tarama", variable=self.bg_scan_var, command=self.toggle_bg_scan, bg="#2b2b2b", fg="#FF5252", selectcolor="#2b2b2b", activebackground="#2b2b2b", activeforeground="#FF5252", font=("Segoe UI", 10, "bold"))
        self.chk_bg_scan.pack(side="left", padx=5)
        
        # UYUMSUZLUK TAKİP BUTONU
        tk.Button(top_frame, text="UYUMSUZLUK TAKİP", command=self.open_monitor_window, 
                  bg="#FF9800", fg="black", font=("Segoe UI", 10, "bold"), relief="flat").pack(side="left", padx=5)

        self.chk_auto = tk.Checkbutton(top_frame, text="Oto. Yenile (5dk)", variable=self.auto_refresh_var, command=self.toggle_auto_refresh, bg="#2b2b2b", fg="white", selectcolor="#2b2b2b", activebackground="#2b2b2b", activeforeground="white", font=("Segoe UI", 10))
        self.chk_auto.pack(side="left", padx=5)

        # Filtreleme Alanı
        filter_frame = tk.Frame(self, bg="#2b2b2b")
        filter_frame.pack(fill="x", padx=10, pady=5)
        
        tk.Label(filter_frame, text="Filtrele:", bg="#2b2b2b", fg="white").pack(side="left")
        self.cb_filter = ttk.Combobox(filter_frame, state="readonly", values=["HEPSİ", "❌ UYUMSUZ", "⚠️ DİKKAT", "✅ UYUMLU"])
        self.cb_filter.current(0)
        self.cb_filter.pack(side="left", padx=5)
        self.cb_filter.bind("<<ComboboxSelected>>", self.apply_filter)
        
        tk.Label(filter_frame, text="Ara:", bg="#2b2b2b", fg="white").pack(side="left", padx=(10, 5))
        self.entry_search = tk.Entry(filter_frame, bg="#1e1e1e", fg="white", insertbackground="white")
        self.entry_search.pack(side="left", padx=5, fill="x", expand=True)
        self.entry_search.bind("<KeyRelease>", self.apply_filter)

        self.lbl_status = tk.Label(top_frame, text="Hazır", bg="#2b2b2b", fg="#aaaaaa")
        self.lbl_status.pack(side="left", padx=10)

        # Treeview Frame (Tablo ve Kaydırma Çubukları için Konteyner)
        tree_frame = tk.Frame(self, bg="#2b2b2b")
        tree_frame.pack(fill="both", expand=True, padx=10, pady=10)

        # Treeview
        cols = ("date", "Türü", "İstasyon", "Uyum", "Bülten")
        self.tree = ttk.Treeview(tree_frame, columns=cols, show="headings")
        self.tree.heading("date", text="Tarih")
        self.tree.heading("Türü", text="Türü")
        self.tree.heading("İstasyon", text="İstasyon")
        self.tree.heading("Uyum", text="TREND UYUM")
        self.tree.heading("Bülten", text="Bülten")
        
        self.tree.column("date", width=100, anchor="center")
        self.tree.column("Türü", width=50, anchor="center")
        self.tree.column("İstasyon", width=60, anchor="center")
        self.tree.column("Uyum", width=200, anchor="center")
        self.tree.column("Bülten", width=900, anchor="w")
        
        sb_y = ttk.Scrollbar(tree_frame, orient="vertical", command=self.tree.yview)
        sb_x = ttk.Scrollbar(tree_frame, orient="horizontal", command=self.tree.xview)
        self.tree.configure(yscrollcommand=sb_y.set, xscrollcommand=sb_x.set)
        
        sb_y.pack(side="right", fill="y")
        sb_x.pack(side="bottom", fill="x")
        self.tree.pack(fill="both", expand=True)
        
        self.tree.tag_configure('UYUMSUZ', background='#D32F2F', foreground='white')
        self.tree.tag_configure('DIKKAT', background='#FFD700', foreground='black')
        self.tree.tag_configure('UYUMLU', foreground='#69F0AE')
        self.tree.tag_configure('SİNOPTİK', foreground='#FFB74D')
        self.tree.tag_configure('METAR', foreground='#4FC3F7')
        self.tree.tag_configure('TAF', background='#3E2723', foreground='#FFAB00', font=("Verdana", 10, "bold"))
        self.tree.tag_configure('oddrow', background='#303f46')

        # Detay Paneli
        self.detail_text = tk.Text(self, height=8, bg="#1e1e1e", fg="white", font=("Consolas", 11), wrap="word")
        self.detail_text.pack(fill="x", padx=10, pady=10)
        
        self.detail_text.tag_config("header", foreground="#aaaaaa", font=("Segoe UI", 10, "bold"))
        self.detail_text.tag_config("content", foreground="white")
        self.detail_text.tag_config("red", foreground="#FF5252")
        self.detail_text.tag_config("yellow", foreground="#FFD700")
        self.detail_text.tag_config("green", foreground="#69F0AE")
        self.detail_text.tag_config("metar_color", foreground="#4FC3F7")
        self.detail_text.tag_config("taf_style", foreground="#FFAB00", font=("Consolas", 11, "bold"))
        self.detail_text.tag_config("default", foreground="white")
        self.detail_text.tag_config("highlight", background="#FFEB3B", foreground="black")
        
        self.tree.bind("<<TreeviewSelect>>", self.on_select)
        self.tree.bind("<Motion>", self.on_tree_motion)
        self.tree.bind("<Leave>", self.hide_tooltip)
        self.tree.bind("<Double-1>", self.open_detail_window)
        self.tree.bind("<Button-3>", self.show_tree_context_menu)
        self.detail_text.bind("<Button-3>", self.show_text_context_menu)

    def toggle_bg_scan(self):
        if self.bg_scan_var.get():
            self.bg_scan_loop()
            messagebox.showinfo("Bilgi", "Arka plan taraması başlatıldı.\nProgram açık olduğu sürece her 10 dakikada bir Türkiye geneli taranacak ve uyumsuzluk varsa uyarı verilecek.")

    def bg_scan_loop(self):
        if self.bg_scan_var.get():
            threading.Thread(target=self.perform_background_scan, daemon=True).start()
            self.after(600000, self.bg_scan_loop) # 10 dakika

    def perform_background_scan(self):
        now = datetime.now(timezone.utc).replace(tzinfo=None)
        s_dt = now - timedelta(hours=48) # 48 Saatlik Tarama
        e_dt = now
        incompatible_list = [] # Popup için (Sadece son 1 saat)
        incompatible_rows = []
        
        for code in TURKEY_STATIONS.keys():
            if not self.bg_scan_var.get(): return
            try:
                lines = RASATLAR.fetch(s_dt, e_dt, station=code, timeout=10)
                if lines:
                    df = process_data(lines, code, "", ref_dt=e_dt)
                    tafs = df[df['Türü'] == 'TAF'].sort_values(by='_dt')
                    metars = df[df['Türü'].isin(['METAR', 'SPECI'])].sort_values(by='_dt')
                    
                    if not metars.empty and not tafs.empty:
                        # Son 48 saatteki TÜM METAR'ları kontrol et
                        for _, last_metar in metars.iterrows():
                            metar_txt = last_metar['Bülten']
                            metar_dt = last_metar['_dt']
                            
                            rel_tafs = tafs[tafs['_dt'] <= metar_dt]
                            if rel_tafs.empty: continue
                            
                            target_taf = rel_tafs.iloc[-1]
                            # 3 saat kuralı (Eski TAF ile eşleşmeyi önle)
                            if (metar_dt - target_taf['_dt']) > timedelta(hours=3): continue

                            taf_txt = target_taf['Bülten']
                            taf_dt = target_taf['_dt']
                            
                            # TAF Zamanı
                            regex_period = r'(?:0[1-9]|[12]\d|3[01])(?:[01]\d|2[0-4])/(?:0[1-9]|[12]\d|3[01])(?:[01]\d|2[0-4])'
                            t_valid = re.search(r'\b' + regex_period + r'\b', taf_txt)
                            taf_zaman = t_valid.group(0) if t_valid else "0000/0000"
                            
                            # FM Grubu Analizi
                            active_taf = taf_txt
                            try:
                                best_change_start = -1
                                change_pattern = r'\bFM(?P<fm>\d{6})\b'
                                for m in re.finditer(change_pattern, taf_txt):
                                    start_dt = None
                                    try:
                                        if m.group('fm'):
                                            time_code = m.group('fm')
                                            day, hour, minute = int(time_code[0:2]), int(time_code[2:4]), int(time_code[4:6])
                                            start_dt = taf_dt.replace(day=day, hour=hour, minute=minute, second=0, microsecond=0)
                                        
                                        if start_dt:
                                            if start_dt.day < taf_dt.day and (taf_dt.day - start_dt.day) > 15:
                                                if start_dt.month == 12: start_dt = start_dt.replace(year=start_dt.year+1, month=1)
                                                else: start_dt = start_dt.replace(month=start_dt.month+1)
                                            elif start_dt.day > taf_dt.day and (start_dt.day - taf_dt.day) > 15:
                                                if start_dt.month == 1: start_dt = start_dt.replace(year=start_dt.year-1, month=12)
                                                else: start_dt = start_dt.replace(month=start_dt.month-1)
                                            
                                            if start_dt <= metar_dt:
                                                best_change_start = max(best_change_start, m.start())
                                    except: continue

                                if best_change_start != -1:
                                    active_taf = taf_txt[best_change_start:]
                            except: pass
                            
                            trend_part = ""
                            tr_m = re.search(r'\b(BECMG|TEMPO|NOSIG)\b', metar_txt)
                            if tr_m: trend_part = metar_txt[tr_m.start():]
                            
                            skor, status_code, reasons = robot.analiz_et(active_taf, metar_txt, trend_part, taf_zaman)
                            
                            # UYUMSUZ veya DİKKAT durumlarını yakala
                            if "UYUMSUZ" in status_code or "DİKKAT" in status_code:
                                detay_str = ""
                                icon = ""
                                
                                if "UYUMSUZ" in status_code:
                                    icon = "❌ UYUMSUZ"
                                    detay_str = "1- UYUMSUZLUK NEDENİ:\n" + "\n".join([f"• {r}" for r in reasons])
                                    detay_str += "\n\n2- TREND KONTROLÜ:\n• Trend ile de uyum sağlanamadı veya Trend yok."
                                    detay_str += "\n\n3- SONUÇ:\n• ❌ UYUMSUZ"
                                    
                                    # Popup uyarısı sadece son 1 saat içindekiler için verilsin (48 saatlik spam olmasın)
                                    if (now - metar_dt) < timedelta(hours=1):
                                        incompatible_list.append(f"{code}: {reasons[0] if reasons else 'Uyumsuz'}")

                                elif "DİKKAT" in status_code:
                                    icon = "⚠️ DİKKAT"
                                    detay_str = "1- UYUMSUZLUK NEDENİ (Ana METAR):\n" + "\n".join([f"• {r}" for r in reasons])
                                    if any("TAF Trend" in r for r in reasons):
                                        detay_str += "\n\n2- TREND KONTROLÜ:\n• ✅ TAF Trendi ile erken uyum (Buffer)."
                                        detay_str += "\n\n3- SONUÇ:\n• ⚠️ DİKKAT (TAF Trendi ile uyumlu)"
                                    else:
                                        detay_str += "\n\n2- TREND KONTROLÜ:\n• ✅ METAR Trendi TAF limitlerine giriyor."
                                        detay_str += "\n\n3- SONUÇ:\n• ⚠️ DİKKAT (METAR Trendi ile uyumlu)"
                                
                                row_data = last_metar.to_dict()
                                row_data["_uyum"] = icon
                                row_data["_detay"] = detay_str
                                row_data["_ref_taf"] = taf_txt
                                incompatible_rows.append(row_data)
            except: pass
            
        if incompatible_list and self.bg_scan_var.get():
            report = "⚠️ ARKA PLAN TARAMA UYARISI ⚠️\n\nAşağıdaki istasyonlarda uyumsuzluk tespit edildi:\n\n" + "\n".join(incompatible_list)
            self.after(0, lambda: messagebox.showwarning("UYUMSUZ RASAT ALARMI", report))
            if incompatible_rows:
                self.after(0, lambda: self.add_background_results(incompatible_rows))
                self.after(0, lambda: self.add_to_monitor(incompatible_rows))

    def add_background_results(self, rows):
        new_df = pd.DataFrame(rows)
        if self.full_df is None or self.full_df.empty:
            self.full_df = new_df
        else:
            self.full_df = pd.concat([self.full_df, new_df]).drop_duplicates(subset=['İstasyon', 'date', 'Bülten']).sort_values(by='_dt', ascending=False)
        
        self.update_tree(self.full_df)

    def add_to_monitor(self, rows):
        """Uyumsuzluk takip listesine veri ekler."""
        new_df = pd.DataFrame(rows)
        if self.incompatible_df is None or self.incompatible_df.empty:
            self.incompatible_df = new_df
        else:
            self.incompatible_df = pd.concat([self.incompatible_df, new_df])
            self.incompatible_df = self.incompatible_df.drop_duplicates(subset=['İstasyon', 'date', 'Bülten'])
        
        # Yeniden eskiye sırala
        if '_dt' in self.incompatible_df.columns:
            self.incompatible_df = self.incompatible_df.sort_values(by='_dt', ascending=False)
            
        if self.monitor_window and tk.Toplevel.winfo_exists(self.monitor_window):
            self.refresh_monitor_tree()

    def open_monitor_window(self):
        """Uyumsuzluk Takip Penceresini Açar."""
        if self.monitor_window is None or not tk.Toplevel.winfo_exists(self.monitor_window):
            self.monitor_window = tk.Toplevel(self)
            self.monitor_window.title("Uyumsuzluk Takip Ekranı")
            self.monitor_window.geometry("1100x600")
            self.monitor_window.configure(bg="#2b2b2b")
            
            lbl = tk.Label(self.monitor_window, text="ARKA PLAN TARAMASI - TESPİT EDİLEN UYUMSUZLUKLAR", bg="#2b2b2b", fg="#FF5252", font=("Segoe UI", 12, "bold"))
            lbl.pack(pady=10)
            
            cols = ("date", "İstasyon", "Uyum", "Bülten", "Detay")
            self.monitor_tree = ttk.Treeview(self.monitor_window, columns=cols, show="headings")
            self.monitor_tree.heading("date", text="Tarih")
            self.monitor_tree.heading("İstasyon", text="İstasyon")
            self.monitor_tree.heading("Uyum", text="Durum")
            self.monitor_tree.heading("Bülten", text="METAR")
            self.monitor_tree.heading("Detay", text="Neden (Özet)")
            
            self.monitor_tree.column("date", width=120, anchor="center")
            self.monitor_tree.column("İstasyon", width=80, anchor="center")
            self.monitor_tree.column("Uyum", width=150, anchor="center")
            self.monitor_tree.column("Bülten", width=400, anchor="w")
            self.monitor_tree.column("Detay", width=300, anchor="w")
            
            sb = ttk.Scrollbar(self.monitor_window, orient="vertical", command=self.monitor_tree.yview)
            self.monitor_tree.configure(yscrollcommand=sb.set)
            sb.pack(side="right", fill="y")
            self.monitor_tree.pack(fill="both", expand=True, padx=10, pady=10)
            
            self.monitor_tree.tag_configure('UYUMSUZ', background='#D32F2F', foreground='white')
            self.monitor_tree.tag_configure('DIKKAT', background='#FFD700', foreground='black')
            
            self.refresh_monitor_tree()
        else:
            self.monitor_window.lift()

    def refresh_monitor_tree(self):
        if self.monitor_tree is None: return
        try:
            self.monitor_tree.delete(*self.monitor_tree.get_children())
            if self.incompatible_df is not None and not self.incompatible_df.empty:
                for _, row in self.incompatible_df.iterrows():
                    # Detay sütunu için kısa özet (İlk satırı al)
                    detay_full = row.get("_detay", "")
                    detay_short = ""
                    if detay_full:
                        lines = detay_full.split('\n')
                        for l in lines:
                            if l.strip().startswith("•"):
                                detay_short = l.strip()
                                break
                    
                    vals = (row.get("date", ""), row.get("İstasyon", ""), row.get("_uyum", ""), row.get("Bülten", ""), detay_short)
                    tag = 'UYUMSUZ' if 'UYUMSUZ' in row.get("_uyum", "") else 'DIKKAT'
                    self.monitor_tree.insert("", "end", values=vals, tags=(tag,))
        except: pass

    def auto_save_monitor_loop(self):
        """Uyumsuzluk listesini her saat başı otomatik kaydeder."""
        if self.incompatible_df is not None and not self.incompatible_df.empty:
            try:
                fname = f"Oto_Kayit_Uyumsuzluk_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx"
                # Sütunları düzenle
                save_df = self.incompatible_df.copy()
                cols_to_export = ["date", "İstasyon", "Bülten", "_uyum", "_detay", "_ref_taf"]
                available_cols = [c for c in cols_to_export if c in save_df.columns]
                save_df = save_df[available_cols]
                save_df.to_excel(fname, index=False)
                print(f"Otomatik kayıt yapıldı: {fname}")
            except Exception as e: print(f"Oto kayıt hatası: {e}")
        
        self.after(3600000, self.auto_save_monitor_loop) # 1 saat (3600000 ms)

    def open_turkey_map(self):
        """Türkiye haritası ve toplu analiz penceresini açar."""
        map_win = tk.Toplevel(self)
        map_win.title("Türkiye Geneli ICAO Analiz Haritası")
        map_win.geometry("1100x650")
        map_win.state('zoomed') # Tam ekran (Maximized)
        map_win.configure(bg="#263238")
        
        # Üst Bilgi
        tk.Label(map_win, text="TÜRKİYE GENELİ METAR/TAF ANALİZİ", bg="#263238", fg="white", font=("Segoe UI", 14, "bold")).pack(pady=10)
        
        # Harita Alanı (Canvas)
        canvas_frame = tk.Frame(map_win, bg="#37474f", bd=2, relief="sunken")
        canvas_frame.pack(fill="both", expand=True, padx=20, pady=10)
        
        canvas = tk.Canvas(canvas_frame, bg="#263238", highlightthickness=0)
        canvas.pack(fill="both", expand=True)
        
        # Alt Kontrol Paneli
        ctrl_frame = tk.Frame(map_win, bg="#263238")
        ctrl_frame.pack(fill="x", padx=20, pady=10)
        
        lbl_map_status = tk.Label(ctrl_frame, text="Analiz için 'TARAMAYI BAŞLAT' butonuna basın. (Son 3 saatlik veriler taranır)", bg="#263238", fg="#B0BEC5", font=("Segoe UI", 10))
        lbl_map_status.pack(side="left")
        
        btn_start = tk.Button(ctrl_frame, text="TARAMAYI BAŞLAT / YENİLE", bg="#00E676", fg="black", font=("Segoe UI", 10, "bold"), width=25)
        btn_start.pack(side="right")
        
        # Auto Refresh (5dk)
        var_map_auto = tk.BooleanVar()
        def map_auto_loop():
            if var_map_auto.get():
                try:
                    if map_win.winfo_exists():
                        if btn_start['state'] != 'disabled':
                            run_scan()
                        map_win.after(300000, map_auto_loop) # 5 dakika
                except: pass
        
        def toggle_map_auto():
            if var_map_auto.get(): map_auto_loop()

        chk_map_auto = tk.Checkbutton(ctrl_frame, text="Oto. Yenile (5dk)", variable=var_map_auto, command=toggle_map_auto, bg="#263238", fg="white", selectcolor="#263238", activebackground="#263238", activeforeground="white", font=("Segoe UI", 10))
        chk_map_auto.pack(side="right", padx=10)
        
        # Filtre Combobox'ı (HATA DÜZELTME: draw_map çağrılmadan önce tanımlanmalı)
        tk.Label(ctrl_frame, text="Filtre:", bg="#263238", fg="white", font=("Segoe UI", 10)).pack(side="left", padx=(20, 5))
        cb_map_filter = ttk.Combobox(ctrl_frame, values=["HEPSİ", "❌ UYUMSUZ", "⚠️ DİKKAT", "✅ UYUMLU"], state="readonly", width=15)
        cb_map_filter.current(0)
        cb_map_filter.pack(side="left")

        # Harita Değişkenleri
        min_lon, max_lon = 25.0, 45.0
        min_lat, max_lat = 35.5, 42.5
        station_items = {}
        station_colors = {} 
        map_results = {} # Detaylı verileri sakla

        # Yanıp Sönme Efekti (Blink)
        map_win.blink_state = False
        def run_blink_effect():
            if not map_win.winfo_exists(): return
            
            map_win.blink_state = not map_win.blink_state
            # Uyumsuzluk rengi (#EF5350) ile Beyaz (#FFFFFF) arasında geçiş yap
            flash_color = "#FFFFFF" if map_win.blink_state else "#EF5350"
            
            for code, color in station_colors.items():
                if color == "#EF5350": # Sadece UYUMSUZ olanları hedefle
                    if code in station_items:
                        try:
                            canvas.itemconfig(station_items[code], fill=flash_color)
                        except: pass
            
            map_win.after(500, run_blink_effect)
        
        run_blink_effect()

        def show_detail(code):
            if code not in map_results:
                messagebox.showinfo(f"{code}", "Detaylı veri bulunamadı.")
                return
            
            data = map_results[code]
            
            top = tk.Toplevel(map_win)
            top.title(f"Detaylı Analiz: {code} - {data['date']}")
            top.geometry("900x700")
            top.configure(bg="#2b2b2b")
            
            tk.Label(top, text=f"METAR - TAF KARŞILAŞTIRMASI ({code})", font=("Segoe UI", 12, "bold"), bg="#2b2b2b", fg="white").pack(pady=10)
            
            # Kopyala Butonu
            btn_frame = tk.Frame(top, bg="#2b2b2b")
            btn_frame.pack(fill="x", padx=10)
            
            main_frame = tk.Frame(top, bg="#2b2b2b")
            main_frame.pack(fill="both", expand=True, padx=10, pady=5)
            
            txt_display = tk.Text(main_frame, bg="#1e1e1e", fg="white", font=("Consolas", 11), relief="flat", wrap="word", padx=10, pady=10)
            sb = ttk.Scrollbar(main_frame, orient="vertical", command=txt_display.yview)
            txt_display.configure(yscrollcommand=sb.set)
            
            sb.pack(side="right", fill="y")
            txt_display.pack(side="left", fill="both", expand=True)
            
            def copy_content():
                try:
                    content = txt_display.get("1.0", tk.END)
                    self.clipboard_clear()
                    self.clipboard_append(content)
                    messagebox.showinfo("Bilgi", "İçerik panoya kopyalandı.")
                except: pass
            
            tk.Button(btn_frame, text="📋 Metni Kopyala", command=copy_content, 
                      bg="#0078D7", fg="white", font=("Segoe UI", 9, "bold")).pack(side="right")
            
            # Tag configurations
            txt_display.tag_config("header", foreground="#aaaaaa", font=("Segoe UI", 10, "bold"))
            txt_display.tag_config("metar", foreground="#4FC3F7", font=("Consolas", 11, "bold"))
            txt_display.tag_config("taf", foreground="#E64A19", font=("Consolas", 11, "bold"))
            txt_display.tag_config("green", foreground="#69F0AE")
            txt_display.tag_config("red", foreground="#FF5252")
            txt_display.tag_config("yellow", foreground="#FFD700")
            txt_display.tag_config("default", foreground="#eceff1")
            
            # İçerik Ekleme
            txt_display.insert("end", "METAR / SPECI:\n", "header")
            txt_display.insert("end", f"{data['metar']}\n\n", "metar")
            
            txt_display.insert("end", "REFERANS TAF:\n", "header")
            txt_display.insert("end", f"{data['taf']}\n\n", "taf")
            
            txt_display.insert("end", "-"*60 + "\n", "header")
            txt_display.insert("end", "ANALİZ DETAYI:\n", "header")
            
            for line in data['detail'].split('\n'):
                tag = "default"
                if "✅" in line or "UYUMLU" in line: tag = "green"
                elif "❌" in line or "UYUMSUZ" in line: tag = "red"
                elif "⚠️" in line or "DİKKAT" in line: tag = "yellow"
                txt_display.insert("end", f"{line}\n", tag)
            
            txt_display.config(state="disabled")
        
        # Filtreleme Değişkeni
        var_filter_incompatible = tk.BooleanVar(value=False)

        def draw_map():
            canvas.delete("all")
            w = canvas.winfo_width()
            h = canvas.winfo_height()
            if w < 100: w = 1060
            if h < 100: h = 500
            
            # Enlem/Boylam oranını koruyarak ölçekleme
            lon_range = max_lon - min_lon
            lat_range = max_lat - min_lat
            
            if w/h > lon_range/lat_range:
                scale = (h * 0.9) / lat_range
            else:
                scale = (w * 0.9) / lon_range
            
            offset_x = (w - (lon_range * scale)) / 2
            offset_y = (h - (lat_range * scale)) / 2
            
            def get_coords(lon, lat):
                x = offset_x + (lon - min_lon) * scale
                y = h - (offset_y + (lat - min_lat) * scale)
                return x, y

            # 1. Türkiye Sınırları (Poligon)
            poly_points = []
            for lon, lat in TURKEY_BORDER:
                px, py = get_coords(lon, lat)
                poly_points.extend([px, py])
            canvas.create_polygon(poly_points, outline="#90A4AE", fill="#37474F", width=2)
            
            # İstasyonları Çiz
            for code, info in TURKEY_STATIONS.items():
                lon, lat = info['lon'], info['lat']
                x, y = get_coords(lon, lat)
                
                # Renk belirle (Varsa kayıtlı rengi kullan)
                fill_color = station_colors.get(code, "#78909C")
                
                # FİLTRELEME KONTROLÜ
                filter_mode = cb_map_filter.get()
                if filter_mode != "HEPSİ":
                    if filter_mode == "❌ UYUMSUZ" and fill_color != "#EF5350": continue
                    elif filter_mode == "⚠️ DİKKAT" and fill_color != "#FFEE58": continue
                    elif filter_mode == "✅ UYUMLU" and fill_color != "#66BB6A": continue
                
                # Marker (Belirgin Daire)
                r = 8 # Yarıçap (Büyütüldü)
                # Dış beyaz halka (Görünürlük için)
                canvas.create_oval(x-r-2, y-r-2, x+r+2, y+r+2, fill="white", outline="")
                # İç renkli daire
                item_id = canvas.create_oval(x-r, y-r, x+r, y+r, fill=fill_color, outline="black", width=1)
                # İstasyon Adı
                canvas.create_text(x, y+r+12, text=code, fill="white", font=("Segoe UI", 9, "bold"))
                
                station_items[code] = item_id
                canvas.tag_bind(item_id, "<Button-1>", lambda e, c=code: show_detail(c))
        
        # Pencere boyutu değişince yeniden çiz
        canvas.bind("<Configure>", lambda e: draw_map())
        
        # Filtre değişince yeniden çiz
        cb_map_filter.bind("<<ComboboxSelected>>", lambda e: draw_map())

        def run_scan():
            if btn_start['state'] == 'disabled': return
            btn_start.config(state="disabled")
            threading.Thread(target=scan_worker, daemon=True).start()

        def scan_worker():
            # Ogimet UTC ile çalışır. Yerel saat farkından dolayı veri gelmemesini önlemek için UTC kullanıyoruz.
            now = datetime.now(timezone.utc).replace(tzinfo=None)
            e_dt = now
            
            incompatible_list = []
            total = len(TURKEY_STATIONS)
            
            for i, code in enumerate(TURKEY_STATIONS.keys(), 1):
                self.after(0, lambda c=code, idx=i: lbl_map_status.config(text=f"Taranıyor ({idx}/{total}): {c}...", fg="#FFD740"))
                
                try:
                    lines = []
                    # TAF bulana kadar geriye dönük tarama (Max 30 saat)
                    for hours_back in range(6, 31, 6):
                        s_dt = now - timedelta(hours=hours_back)
                        
                        lines = RASATLAR.fetch(s_dt, e_dt, station=code, timeout=10)
                        if lines:
                            # TAF var mı kontrol et
                            df_check = process_data(lines, code, "", ref_dt=e_dt)
                            if not df_check[df_check['Türü'] == 'TAF'].empty:
                                break # TAF bulundu

                    color = "#455A64" # Veri yok (Gri)
                    status_msg = "Veri Yok"
                    detail_data = None
                    
                    if lines:
                        df = process_data(lines, code, "", ref_dt=e_dt)
                        tafs = df[df['Türü'] == 'TAF'].sort_values(by='_dt')
                        metars = df[df['Türü'].isin(['METAR', 'SPECI'])].sort_values(by='_dt')
                        
                        if not metars.empty and not tafs.empty:
                            last_metar = metars.iloc[-1]
                            metar_txt = last_metar['Bülten']
                            
                            # İlgili TAF'ı bul
                            rel_tafs = tafs[tafs['_dt'] <= last_metar['_dt']]
                            if not rel_tafs.empty:
                                target_taf = rel_tafs.iloc[-1]
                                taf_txt = target_taf['Bülten']
                                
                                # Analiz Et
                                trend_part = ""
                                tr_m = re.search(r'\b(BECMG|TEMPO|NOSIG)\b', metar_txt)
                                if tr_m: trend_part = metar_txt[tr_m.start():]
                                
                                skor, status_code, reasons = robot.analiz_et(taf_txt, metar_txt, trend_part)
                                
                                # Detay Metni Oluştur
                                detay_str = ""
                                if "UYUMSUZ" in status_code:
                                    color = "#EF5350" # Kırmızı
                                    incompatible_list.append(f"❌ {code}: {reasons[0] if reasons else 'Uyumsuz'}")
                                    status_msg = f"UYUMSUZ\n{reasons[0] if reasons else ''}"
                                    detay_str = "1- UYUMSUZLUK NEDENİ:\n" + "\n".join([f"• {r}" for r in reasons])
                                    detay_str += "\n\n2- TREND KONTROLÜ:\n• Trend ile de uyum sağlanamadı veya Trend yok."
                                    detay_str += "\n\n3- SONUÇ:\n• ❌ UYUMSUZ"
                                elif "DİKKAT" in status_code:
                                    color = "#FFEE58" # Sarı
                                    status_msg = "DİKKAT (Trend ile uyumlu)"
                                    detay_str = "1- UYUMSUZLUK NEDENİ (Ana METAR):\n" + "\n".join([f"• {r}" for r in reasons])
                                    detay_str += "\n\n2- TREND KONTROLÜ:\n• ✅ METAR Trendi TAF limitlerine giriyor."
                                    detay_str += "\n\n3- SONUÇ:\n• ⚠️ DİKKAT (Trend ile uyumlu)"
                                else:
                                    color = "#66BB6A" # Yeşil
                                    status_msg = "UYUMLU"
                                    detay_str = "1- DURUM:\n• TAF limitleri dahilinde."
                                    if "Trend" in status_code:
                                        detay_str += " (TAF Trendi ile)"
                                    detay_str += "\n\n3- SONUÇ:\n• ✅ UYUMLU"
                                
                                detail_data = {
                                    'metar': metar_txt,
                                    'taf': taf_txt,
                                    'detail': detay_str,
                                    'date': last_metar['date']
                                }
                    
                    # Haritayı Güncelle
                    self.after(0, lambda c=code, col=color, m=status_msg, d=detail_data: update_station_ui(c, col, m, d))
                    
                except Exception as e: print(f"Scan error {code}: {e}")
            
            self.after(0, lambda: finalize_scan(incompatible_list))

        def update_station_ui(code, color, msg, detail_data=None):
            station_colors[code] = color
            if detail_data:
                map_results[code] = detail_data
            
            # Marker güncelle (Eğer haritada varsa)
            if code in station_items:
                canvas.itemconfig(station_items[code], fill=color)
            else:
                # Eğer filtre nedeniyle gizliyse, draw_map çağrıldığında doğru renkle çizilecek
                pass

        def finalize_scan(incompatible_list):
            btn_start.config(state="normal")
            lbl_map_status.config(text="Tarama Tamamlandı.", fg="#69F0AE")
            
            if incompatible_list:
                report = "UYUMSUZ RASATLAR TESPİT EDİLDİ:\n\n" + "\n".join(incompatible_list)
                messagebox.showwarning("Tarama Sonucu", report)

        btn_start.config(command=run_scan)
        
        # Otomatik Başlat (Pencere açıldıktan kısa süre sonra)
        map_win.after(500, run_scan)

    def toggle_auto_refresh(self):
        if self.auto_refresh_var.get():
            self.auto_refresh_loop()
        elif self.refresh_job:
            self.after_cancel(self.refresh_job)
            self.refresh_job = None

    def auto_refresh_loop(self):
        if self.auto_refresh_var.get():
            self.start_process()
            self.refresh_job = self.after(300000, self.auto_refresh_loop)

    def start_process(self):
        self.save_config()
        st = self.ent_station.get().strip().upper()
        wmo = self.ent_wmo.get().strip()
        d_start = self.ent_start.get_date()
        d_end = self.ent_end.get_date()
        
        h_start = int(self.cb_start_hour.get())
        h_end = int(self.cb_end_hour.get())
        
        s_dt = datetime.combine(d_start, datetime.min.time().replace(hour=h_start))
        e_dt = datetime.combine(d_end, datetime.min.time().replace(hour=h_end, minute=59))
        
        self.lbl_status.config(text="Veriler çekiliyor...", fg="#FFD740")
        threading.Thread(target=self.worker, args=(st, wmo, s_dt, e_dt), daemon=True).start()

    def worker(self, st, wmo, s_dt, e_dt):
        try:
            all_lines = []
            curr_end = e_dt
            
            # Ogimet genellikle 30-31 günlük veri verir. Uzun aralıklar for döngü kuruyoruz.
            while curr_end > s_dt:
                curr_start = curr_end - timedelta(days=30)
                if curr_start < s_dt:
                    curr_start = s_dt
                
                self.after(0, lambda s=curr_start, e=curr_end: self.lbl_status.config(text=f"Çekiliyor: {s.strftime('%d.%m.%Y')} - {e.strftime('%d.%m.%Y')}", fg="#FFD740"))
                
                try:
                    chunk_lines = RASATLAR.fetch(curr_start, curr_end, station=st, wmo_id=wmo)
                    if chunk_lines:
                        all_lines.extend(chunk_lines)
                except Exception as e:
                    print(f"Veri parça hatası: {e}")
                
                curr_end = curr_start - timedelta(seconds=1)
                time.sleep(0.5) # Sunucuyu yormamak for bekleme

            if not all_lines:
                self.after(0, lambda: messagebox.showwarning("Veri Bulunamadı", f"Sorgulanan Tarih: {s_dt.strftime('%d.%m.%Y')}\nİstasyon: {st}\n\nOlası Sebepler:\n1. Ogimet sunucusu yanıt vermiyor veya boş dönüyor.\n2. Bu tarihte istasyon veri göndermemiş.\n3. İnternet bağlantısı sorunu.\n\nLütfen terminal penceresindeki DEBUG çıktılarını kontrol edin."))
                self.after(0, lambda: self.lbl_status.config(text="Veri yok", fg="white"))
                return

            try:
                df = process_data(all_lines, st, wmo, ref_dt=e_dt)
            except Exception as e:
                raise Exception(f"Veri işleme hatası: {e}")
            
            # ANALİZ
            df["_uyum"] = ""
            df["_detay"] = ""
            df["_ref_taf"] = ""
            
            try:
                tafs = df[df['Türü'] == 'TAF'].sort_values(by='_dt')
                
                if not tafs.empty:
                    # Ardışık Exception takibi for sayaçlar
                    last_taf_text = None
                    consecutive_counts = {"Rüzgar": 0, "Görüş": 0, "ceil": 0}
                    
                    # Kronolojik sıra with işle (Eskiden yeniye)
                    # df.sort_values with sıralı iterasyon yapıyoruz
                    for idx, row in df.sort_values(by='_dt', ascending=True).iterrows():
                        if row['Türü'] in ['METAR', 'SPECI']:
                            metar_dt = row['_dt']
                            
                            # 1. METAR zamanından önce or aynı zamanda yayınlanmış TAF'ları al
                            relevant_tafs = tafs[tafs['_dt'] <= metar_dt]
                            
                            target_row = None
                            
                            # 2. En son yayınlanan TAF'ı seç (passerlilik süresi kontrolü iptal)
                            if not relevant_tafs.empty:
                                target_row = relevant_tafs.iloc[-1]
                            
                            if target_row is not None:
                                last_taf = target_row['Bülten']
                                taf_dt = target_row['_dt']
                                
                                # TAF değiştiyse sayaçları sıfırla
                                if last_taf != last_taf_text:
                                    consecutive_counts = {k: 0 for k in consecutive_counts}
                                    last_taf_text = last_taf
                                
                                # 3 time KURALI: TAF yayınlandıktan sonraki 3 time içindeki rasatlar dikkate alınır.
                                time_diff = metar_dt - taf_dt
                                if time_diff > timedelta(hours=3):
                                    continue

                                print(f"🔍 DENETİM: METAR {metar_dt.strftime('%d/%H:%M')} <--- TAF {target_row['_dt'].strftime('%d/%H:%M')}")

                                df.at[idx, "_ref_taf"] = last_taf
                                try:
                                    # TAF Zamanını bul (Örn: 0412/0512)
                                    # Regex geliştirildi: Gün (01-31) and time (00-24) formatına uygunluk kontrolü. BECMG/TEMPO for de passerli.
                                    regex_period = r'(?:0[1-9]|[12]\d|3[01])(?:[01]\d|2[0-4])/(?:0[1-9]|[12]\d|3[01])(?:[01]\d|2[0-4])'
                                    
                                    t_valid = re.search(r'\b' + regex_period + r'\b', last_taf)
                                    taf_zaman = t_valid.group(0) if t_valid else "0000/0000"
                                    
                                    # FM Gruplarını dikkate alarak aktif splitümü seç
                                    active_taf = last_taf
                                    # FM, BECMG and TEMPO gruplarını dikkate alarak aktif splitümü seç
                                    try:
                                        best_change_start = -1
                                        taf_dt = target_row['_dt']

                                        # Sadece FM grupları ana TAF'ı sıfırlar. BECMG/TEMPO trend olarak işlenir.
                                        change_pattern = r'\bFM(?P<fm>\d{6})\b'
                                        
                                        for m in re.finditer(change_pattern, last_taf):
                                            start_dt = None
                                            try:
                                                if m.group('fm'):
                                                    time_code = m.group('fm')
                                                    day, hour, minute = int(time_code[0:2]), int(time_code[2:4]), int(time_code[4:6])
                                                    start_dt = taf_dt.replace(day=day, hour=hour, minute=minute, second=0, microsecond=0)
                                                
                                                if start_dt:
                                                    # Ay passişi kontrolü
                                                    if start_dt.day < taf_dt.day and (taf_dt.day - start_dt.day) > 15:
                                                        if start_dt.month == 12: start_dt = start_dt.replace(year=start_dt.year+1, month=1)
                                                        else: start_dt = start_dt.replace(month=start_dt.month+1)
                                                    elif start_dt.day > taf_dt.day and (start_dt.day - taf_dt.day) > 15:
                                                        if start_dt.month == 1: start_dt = start_dt.replace(year=start_dt.year-1, month=12)
                                                        else: start_dt = start_dt.replace(month=start_dt.month-1)
                                                    
                                                    if start_dt <= metar_dt:
                                                        best_change_start = max(best_change_start, m.start())
                                            except (ValueError, IndexError): continue

                                        if best_change_start != -1:
                                            active_taf = last_taf[best_change_start:]
                                    except Exception as e:
                                        print(f"Değişim Grubu Analiz Hatası: {e}")

                                    # Trend kısmını ayır
                                    trend_part = ""
                                    tr_m = re.search(r'\b(BECMG|TEMPO|NOSIG)\b', row['Bülten'])
                                    if tr_m: trend_part = row['Bülten'][tr_m.start():]

                                    skor, status_code, reasons = robot.analiz_et(active_taf, row['Bülten'], trend_part, taf_zaman)
                                    
                                    # --- ARDIŞIK Exception KONTROLÜ ---
                                    current_cats = set()
                                    for r in reasons:
                                        if "Rüzgar" in r: current_cats.add("Rüzgar")
                                        if "Görüş" in r: current_cats.add("Görüş")
                                        if "ceil" in r or "Dikey Görüş" in r or "Bulut" in r: current_cats.add("ceil")
                                    
                                    amd_msgs = []
                                    for cat in consecutive_counts:
                                        if cat in current_cats:
                                            consecutive_counts[cat] += 1
                                            if consecutive_counts[cat] >= 3:
                                                amd_msgs.append(f"{cat} ({consecutive_counts[cat]}. kez)")
                                        else:
                                            consecutive_counts[cat] = 0
                                    # -----------------------------

                                    icon = ""
                                    if "UYUMSUZ" in status_code: icon = "❌ UYUMSUZ"
                                    elif "DİKKAT" in status_code: icon = "⚠️ DİKKAT"
                                    elif "UYUMLU" in status_code: icon = "✅ UYUMLU"
                                    
                                    df.at[idx, "_uyum"] = icon
                                    
                                    # Detaylı Açıklama Metni Oluşturma
                                    detay_str = ""
                                    if "UYUMSUZ" in status_code:
                                        detay_str = "1- UYUMSUZLUK NEDENİ:\n" + "\n".join([f"• {r}" for r in reasons])
                                        detay_str += "\n\n2- TREND KONTROLÜ:\n• Trend ile de uyum sağlanamadı veya Trend yok."
                                        detay_str += "\n\n3- SONUÇ:\n• ❌ UYUMSUZ"
                                    elif "DİKKAT" in status_code:
                                        detay_str = "1- UYUMSUZLUK NEDENİ (Ana METAR):\n" + "\n".join([f"• {r}" for r in reasons])
                                        if any("TAF Trend" in r for r in reasons):
                                            detay_str += "\n\n2- TREND KONTROLÜ:\n• ✅ TAF Trendi ile erken uyum (Buffer)."
                                            detay_str += "\n\n3- SONUÇ:\n• ⚠️ DİKKAT (TAF Trendi ile uyumlu)"
                                        else:
                                            detay_str += "\n\n2- TREND KONTROLÜ:\n• ✅ METAR Trendi TAF limitlerine giriyor."
                                            detay_str += "\n\n3- SONUÇ:\n• ⚠️ DİKKAT (METAR Trendi ile uyumlu)"
                                    elif "UYUMLU" in status_code:
                                        detay_str = "1- DURUM:\n• TAF limitleri dahilinde."
                                        if "Trend" in status_code:
                                            detay_str += " (TAF Trendi ile)"
                                        detay_str += "\n\n3- SONUÇ:\n• ✅ UYUMLU"
                                    
                                    # TAVSİYE EKLE
                                    if amd_msgs:
                                        detay_str += f"\n\n👉 KRİTİK TAVSİYE:\n• TAF AMD YAYINLANMALI!\n  Aynı sapma 3 ve üzerinde tekrarlandı: {', '.join(amd_msgs)}"
                                    
                                    df.at[idx, "_detay"] = detay_str
                                except Exception as e: print(f"Analiz satır hatası: {e}")
            except Exception as e:
                print(f"Genel analiz hatası: {e}")

            self.after(0, lambda: self.update_tree(df))
        except Exception as e:
            self.after(0, lambda: messagebox.showerror("Exception", str(e)))

    def apply_filter(self, event=None):
        if self.full_df is None: return
        choice = self.cb_filter.get()
        search = self.entry_search.get().lower()
        df_show = self.full_df.copy()
        
        if choice == "❌ UYUMSUZ": df_show = df_show[df_show["_uyum"].str.contains("UYUMSUZ", na=False)]
        elif choice == "⚠️ DİKKAT": df_show = df_show[df_show["_uyum"].str.contains("DİKKAT", na=False)]
        elif choice == "✅ UYUMLU": df_show = df_show[df_show["_uyum"].str.contains("UYUMLU", na=False)]
        
        if search: df_show = df_show[df_show["Bülten"].str.lower().str.contains(search, na=False)]
        self.render_tree(df_show)

    def update_tree(self, df):
        self.full_df = df
        self.apply_filter()

    def show_tree_context_menu(self, event):
        menu = tk.Menu(self, tearoff=0)
        menu.add_command(label="Satırı copy", command=self.copy_tree_selection)
        menu.tk_popup(event.x_root, event.y_root)

    def copy_tree_selection(self):
        selected = self.tree.selection()
        if not selected: return
        res = []
        for item in selected:
            vals = self.tree.item(item, "values")
            res.append(" | ".join(map(str, vals)))
        self.clipboard_clear()
        self.clipboard_append("\n".join(res))

    def show_text_context_menu(self, event):
        menu = tk.Menu(self, tearoff=0)
        menu.add_command(label="copy", command=self.copy_text_selection)
        menu.add_command(label="Tümünü Seç", command=self.select_all_text)
        menu.tk_popup(event.x_root, event.y_root)

    def copy_text_selection(self):
        try:
            sel = self.detail_text.get("sel.first", "sel.last")
            self.clipboard_clear()
            self.clipboard_append(sel)
        except: pass

    def select_all_text(self):
        self.detail_text.tag_add("sel", "1.0", "end")

    def render_tree(self, df):
        self.tree_tooltips.clear()
        self.tree.delete(*self.tree.get_children())
        for i, (_, row) in enumerate(df.iterrows()):
            vals = (row["date"], row["Türü"], row["İstasyon"], row["_uyum"], row["Bülten"])
            tag = ""
            if "UYUMSUZ" in row["_uyum"]: tag = "UYUMSUZ"
            elif "DİKKAT" in row["_uyum"]: tag = "DIKKAT"
            elif "UYUMLU" in row["_uyum"]: tag = "UYUMLU"
            elif row["Türü"] == "SİNOPTİK": tag = "SİNOPTİK"
            elif row["Türü"] == "TAF": tag = "TAF"
            elif row["Türü"] in ["METAR", "SPECI"]: tag = "METAR"
            
            row_tags = []
            if i % 2 == 1: row_tags.append('oddrow')
            if tag: row_tags.append(tag)
            
            item_id = self.tree.insert("", "end", values=vals, tags=tuple(row_tags))
            
            if row["_detay"]:
                color = "#eceff1"
                if "UYUMSUZ" in row["_uyum"]: color = "#FF5252"
                elif "DİKKAT" in row["_uyum"]: color = "#FFD700"
                elif "UYUMLU" in row["_uyum"]: color = "#69F0AE"
                self.tree_tooltips[item_id] = (f"Analiz Detayı:\n{row['_detay']}", color)
                
        self.lbl_status.config(text=f"Gösterilen: {len(df)} kayıt", fg="#69F0AE")
        self.adjust_column_widths(df)

    def on_tree_motion(self, event):
        item_id = self.tree.identify_row(event.y)
        if item_id == self.last_tooltip_item: return
        self.last_tooltip_item = item_id
        self.hide_tooltip()
        if item_id and item_id in self.tree_tooltips:
            val = self.tree_tooltips[item_id]
            if isinstance(val, tuple): text, color = val
            else: text, color = val, "#eceff1"
            self.show_tooltip(event.x_root, event.y_root, text, color)

    def show_tooltip(self, x, y, text, text_color="#eceff1"):
        self.tooltip_window = tk.Toplevel(self)
        self.tooltip_window.wm_overrideredirect(True)
        self.tooltip_window.withdraw() 
        bg_color = "#263238" 
        fg_color = text_color 
        border_color = "#FF5252" if "UYUMSUZ" in text or "KRİTİK" in text or "YANLIŞ" in text else "#80cbc4"
        
        frame = tk.Frame(self.tooltip_window, bg=bg_color, highlightbackground=border_color, highlightthickness=2)
        frame.pack(fill="both", expand=True)
        
        # Renkli text for Text widget kullanımı
        lines = text.split('\n')
        w = min(max([len(line) for line in lines] + [40]), 90) + 2
        h = len(lines) + 2
        txt_widget = tk.Text(frame, width=w, height=h, bg=bg_color, fg=fg_color, font=("Consolas", 10), relief="flat", wrap="word")
        txt_widget.pack(padx=10, pady=8)
        
        txt_widget.tag_config("green", foreground="#00E676")
        txt_widget.tag_config("red", foreground="#FF5252")
        txt_widget.tag_config("yellow", foreground="#FFD700")
        txt_widget.tag_config("default", foreground=fg_color)
        
        for line in text.split('\n'):
            tag = "default"
            if "✅" in line or "UYUMLU" in line: tag = "green"
            elif "❌" in line or "UYUMSUZ" in line: tag = "red"
            elif "⚠️" in line or "DİKKAT" in line: tag = "yellow"
            txt_widget.insert("end", line + "\n", tag)
            
        txt_widget.config(state="disabled")
        
        self.tooltip_window.update_idletasks()
        width = self.tooltip_window.winfo_reqwidth()
        height = self.tooltip_window.winfo_reqheight()
        
        pos_x = x + 15
        pos_y = y + 10
        if pos_x + width > self.tooltip_window.winfo_screenwidth(): pos_x = x - width - 15
        if pos_y + height > self.tooltip_window.winfo_screenheight(): pos_y = y - height - 10
        if pos_x < 0: pos_x = 0
        if pos_y < 0: pos_y = 0
        
        self.tooltip_window.wm_geometry(f"+{pos_x}+{pos_y}")
        self.tooltip_window.deiconify()

    def hide_tooltip(self, event=None):
        if self.tooltip_window:
            self.tooltip_window.destroy()
            self.tooltip_window = None
        if event: self.last_tooltip_item = None

    def open_detail_window(self, event=None):
        sel = self.tree.selection()
        if not sel: return
        item = self.tree.item(sel[0])
        vals = item['values']
        
        date = vals[0]
        bulten = vals[4]
        
        if self.full_df is not None:
            row = self.full_df[(self.full_df['date'] == date) & (self.full_df['Bülten'] == bulten)]
            if not row.empty:
                row = row.iloc[0]

                ref_taf = row.get('_ref_taf', '')
                detay = row.get('_detay', '')
                
                top = tk.Toplevel(self)
                top.title(f"Detaylı Analiz: {date}")
                top.geometry("900x700")
                top.configure(bg="#2b2b2b")
                
                tk.Label(top, text=f"METAR - TAF KARŞILAŞTIRMASI ({STATION})", font=("Segoe UI", 12, "bold"), bg="#2b2b2b", fg="white").pack(pady=10)
                
                # Kopyala Butonu Alanı
                btn_frame = tk.Frame(top, bg="#2b2b2b")
                btn_frame.pack(fill="x", padx=10)
                
                # Tek Bölümde Birleşik Görünüm
                main_frame = tk.Frame(top, bg="#2b2b2b")
                main_frame.pack(fill="both", expand=True, padx=10, pady=5)
                
                txt_display = tk.Text(main_frame, bg="#1e1e1e", fg="white", font=("Consolas", 11), relief="flat", wrap="word", padx=10, pady=10)
                sb = ttk.Scrollbar(main_frame, orient="vertical", command=txt_display.yview)
                txt_display.configure(yscrollcommand=sb.set)
                
                sb.pack(side="right", fill="y")
                txt_display.pack(side="left", fill="both", expand=True)
                
                def copy_content():
                    try:
                        content = txt_display.get("1.0", tk.END)
                        self.clipboard_clear()
                        self.clipboard_append(content)
                        messagebox.showinfo("Bilgi", "İçerik panoya kopyalandı.")
                    except: pass
                
                tk.Button(btn_frame, text="📋 Metni Kopyala", command=copy_content, 
                          bg="#0078D7", fg="white", font=("Segoe UI", 9, "bold")).pack(side="right")
                
                # Tag configurations
                txt_display.tag_config("header", foreground="#aaaaaa", font=("Segoe UI", 10, "bold"))
                txt_display.tag_config("metar", foreground="#4FC3F7", font=("Consolas", 11, "bold"))
                txt_display.tag_config("taf", foreground="#E64A19", font=("Consolas", 11, "bold"))
                txt_display.tag_config("green", foreground="#69F0AE")
                txt_display.tag_config("red", foreground="#FF5252")
                txt_display.tag_config("yellow", foreground="#FFD700")
                txt_display.tag_config("default", foreground="#eceff1")
                txt_display.tag_config("highlight", background="#FFEB3B", foreground="black")
                
                # İçerik Ekleme
                if row['Türü'] == 'TAF':
                    txt_display.insert("end", "SEÇİLEN TAF:\n", "header")
                    txt_display.insert("end", f"{bulten}\n\n", "taf")
                else:
                    txt_display.insert("end", "METAR / SPECI:\n", "header")
                    txt_display.insert("end", f"{bulten}\n\n", "metar")
                    
                    txt_display.insert("end", "REFERANS TAF:\n", "header")
                    taf_text = ref_taf if ref_taf else "Uygun TAF bulunamadı."
                    txt_display.insert("end", f"{taf_text}\n\n", "taf")
                    
                    txt_display.insert("end", "-"*60 + "\n", "header")
                    txt_display.insert("end", "ANALİZ DETAYI:\n", "header")
                    
                    for line in detay.split('\n'):
                        tag = "default"
                        if "✅" in line or "UYUMLU" in line: tag = "green"
                        elif "❌" in line or "UYUMSUZ" in line: tag = "red"
                        elif "⚠️" in line or "DİKKAT" in line: tag = "yellow"
                        txt_display.insert("end", f"{line}\n", tag)
                        
                        # Değer Vurgulama Mantığı (Highlight)
                        if "Beklenen:" in line and "METAR:" in line:
                            try:
                                m = re.search(r'Beklenen:(.*?) vs METAR:(.*?)\)', line)
                                if m:
                                    vals = [m.group(1).strip(), m.group(2).strip()]
                                    for v in vals:
                                        if v.endswith('m') and v[:-1].isdigit(): v = v[:-1]
                                        if not v: continue
                                        
                                        start = "1.0"
                                        while True:
                                            pos = txt_display.search(v, start, stopindex=tk.END)
                                            if not pos: break
                                            end = f"{pos}+{len(v)}c"
                                            txt_display.tag_add("highlight", pos, end)
                                            start = end
                            except: pass
                
                txt_display.config(state="disabled")
    def adjust_column_widths(self, df):
        try:
            font = tkfont.Font(family="Segoe UI", size=10)
            header_font = tkfont.Font(family="Segoe UI", size=11, weight="bold")
            
            col_map = {
                "date": "date",
                "Türü": "Türü",
                "İstasyon": "İstasyon",
                "Uyum": "_uyum",
                # "Bülten": "Bülten"
            }
            
            # Sabit sütunları içeriğe göre ayarla (Bülten hariç)
            for col_id in ["date", "Türü", "İstasyon", "Uyum"]:
                df_col = col_map.get(col_id)
                heading_text = self.tree.heading(col_id, "text")
                max_w = header_font.measure(heading_text) + 25
                
                if df_col and df_col in df.columns:
                    vals = df[df_col].fillna("").astype(str).unique()
                    if len(vals) > 20: vals = sorted(vals, key=len, reverse=True)[:20]
                    for v in vals:
                        w = font.measure(v) + 20
                        if w > max_w: max_w = w
                
                if col_id == "Uyum":
                    max_w = max(max_w, 150)
                
                self.tree.column(col_id, width=max_w, stretch=False)
            
            # Bülten sütunu sabit genişlik
            self.tree.column("Bülten", width=900, stretch=False)
            
        except Exception as e: print(f"Resize error: {e}")

    def on_select(self, event):
        sel = self.tree.selection()
        if not sel: return
        item = self.tree.item(sel[0])
        vals = item['values']
        # Detayları bul
        bulten = vals[4]
        date = vals[0]
        
        # DataFrame'den detayı çek
        if self.full_df is not None:
            row = self.full_df[(self.full_df['date'] == date) & (self.full_df['Bülten'] == bulten)]
            if not row.empty:
                row_data = row.iloc[0]
                detay = row_data.get('_detay', '')
                ref_taf = row_data.get('_ref_taf', '')
                
                self.detail_text.config(state="normal")
                self.detail_text.delete("1.0", tk.END)
                
                # Bülten Tipi Kontrolü
                if row_data['Türü'] == 'TAF':
                    self.detail_text.insert(tk.END, "TAF:\n", "header")
                    self.detail_text.insert(tk.END, f"{bulten}\n", "taf_style")
                else:
                    self.detail_text.insert(tk.END, "METAR / SPECI:\n", "header")
                    self.detail_text.insert(tk.END, f"{bulten}\n", "metar_color")
                
                # TAF
                if ref_taf:
                    self.detail_text.insert(tk.END, f"\nİLGİLİ TAF:\n{ref_taf}\n", "taf_style")
                
                # ANALİZ
                if detay:
                    self.detail_text.insert(tk.END, "\nANALİZ DETAYI:\n", "header")

                    reason_tag = "content"
                    uyum_durumu = row_data.get('_uyum', '')
                    if "UYUMSUZ" in uyum_durumu: reason_tag = "red"
                    elif "DİKKAT" in uyum_durumu: reason_tag = "yellow"

                    for line in detay.split('\n'):
                        tag = "default"
                        if "✅" in line or "UYUMLU" in line: tag = "green"
                        elif "❌" in line or "UYUMSUZ" in line: tag = "red"
                        elif "⚠️" in line or "DİKKAT" in line: tag = "yellow"
                        elif line.strip().startswith("•"): tag = reason_tag
                        self.detail_text.insert(tk.END, f"{line}\n", tag)
                        
                        # Değer Vurgulama Mantığı (Highlight)
                        if "Beklenen:" in line and "METAR:" in line:
                            try:
                                m = re.search(r'Beklenen:(.*?) vs METAR:(.*?)\)', line)
                                if m:
                                    vals = [m.group(1).strip(), m.group(2).strip()]
                                    for v in vals:
                                        if v.endswith('m') and v[:-1].isdigit(): v = v[:-1]
                                        if not v: continue
                                        
                                        start = "1.0"
                                        while True:
                                            pos = self.detail_text.search(v, start, stopindex=tk.END)
                                            if not pos: break
                                            end = f"{pos}+{len(v)}c"
                                            self.detail_text.tag_add("highlight", pos, end)
                                            start = end
                            except: pass
                
                self.detail_text.config(state="disabled")


    def export_to_excel(self):
        if self.full_df is None or self.full_df.empty:
            messagebox.showwarning("Uyarı", "Dışa aktarılacak veri yok.\n\nLütfen önce 'VERİ ÇEK' butonuna basınız.\nEğer veri gelmiyorsa tarih aralığını ve sistem saatinizi kontrol ediniz.")
            return
            
        try:
            file_path = filedialog.asksaveasfilename(
                defaultextension=".xlsx",
                filetypes=[("Excel Dosyası", "*.xlsx")],
                title="Analiz Raporunu Kaydet"
            )
            
            if not file_path: return
            
            export_df = self.full_df.copy()
            
            # Sütunları düzenle
            cols_to_export = ["date", "Türü", "İstasyon", "Bülten", "_uyum", "_detay", "_ref_taf"]
            available_cols = [c for c in cols_to_export if c in export_df.columns]
            export_df = export_df[available_cols]
            
            rename_map = {"_uyum": "Uyum Durumu", "_detay": "Analiz Detayı", "_ref_taf": "Referans TAF"}
            export_df.rename(columns=rename_map, inplace=True)
            
            with pd.ExcelWriter(file_path, engine='openpyxl') as writer:
                export_df.to_excel(writer, sheet_name='Tüm Veriler', index=False)
                if "Uyum Durumu" in export_df.columns:
                    uyumsuz_df = export_df[export_df["Uyum Durumu"].str.contains("UYUMSUZ|DİKKAT", na=False)].copy()
                    if not uyumsuz_df.empty:
                        # Detayları sütunlara ayır
                        if "Analiz Detayı" in uyumsuz_df.columns:
                            split_data = uyumsuz_df["Analiz Detayı"].str.split('\n', expand=True)
                            split_data.columns = [f"Neden {i+1}" for i in range(split_data.shape[1])]
                            uyumsuz_df = pd.concat([uyumsuz_df, split_data], axis=1)
                        
                        uyumsuz_df.to_excel(writer, sheet_name='Uyumsuzluklar', index=False)
                
                # Metin Kaydırma (Wrap Text) - Tüm Sayfalar İçin
                try:
                    from openpyxl.styles import Alignment, Font
                    from openpyxl.utils import get_column_letter
                    
                    for sheet_name in writer.sheets:
                        ws = writer.sheets[sheet_name]
                        
                        # Başlıkları Kalın Yap
                        for cell in ws[1]:
                            cell.font = Font(bold=True)
                        
                        # Sütun Genişliklerini Ayarla (Wrap Text için kritik)
                        for i, col in enumerate(ws.columns, 1):
                            col_letter = get_column_letter(i)
                            header = str(col[0].value)
                            
                            if "Bülten" in header or "Analiz" in header or "TAF" in header:
                                ws.column_dimensions[col_letter].width = 60 # Genişlik sabit, metin aşağı kayar
                            elif "Tarih" in header or "date" in header:
                                ws.column_dimensions[col_letter].width = 18
                            else:
                                ws.column_dimensions[col_letter].width = 14
                        
                        for row in ws.iter_rows():
                            for cell in row:
                                cell.alignment = Alignment(wrap_text=True, vertical='top')
                except Exception as e: print(f"Excel stil hatası: {e}")

            messagebox.showinfo("Başarılı", f"Rapor kaydedildi:\n{file_path}")
        except Exception as e:
            messagebox.showerror("Exception", f"Dışa aktarma hatası: {e}")

if __name__ == "__main__":
    app = App()
    app.mainloop()