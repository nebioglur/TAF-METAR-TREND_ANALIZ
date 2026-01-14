# -*- coding: utf-8 -*-
"""
OGIMET ICAO ANALİZ - WEB ARAYÜZÜ (Streamlit)
Bu dosya, masaüstü uygulamasının web versiyonudur.
Çalıştırmak için terminale: streamlit run ogimet_webapp.py
"""

import streamlit as st
import pandas as pd
import re
from datetime import datetime, timedelta, timezone
import RASATLAR
import TAF_METAR_TREND
import io
import plotly.express as px

# Sayfa Ayarları
st.set_page_config(
    page_title="OGIMET ICAO ANALİZ",
    page_icon="✈️",
    layout="wide",
    initial_sidebar_state="expanded"
)

# Robot Modülünü Başlat
@st.cache_resource
def get_robot():
    return TAF_METAR_TREND.HavacilikRobotModulu()

robot = get_robot()

def process_data(lines, station_code, wmo_id):
    """Ham metin verilerini DataFrame'e dönüştürür."""
    data = []
    current_record = None

    for line in lines:
        line = line.strip()
        if not line: continue
        
        # Ogimet HTML kalıntılarını temizle
        if "=" in line and "<" in line:
            line = re.sub(r'=\s*<.*$', '=', line)
        
        parts = line.split()
        is_start = False
        
        if len(parts) > 0:
            if parts[0].isdigit() and len(parts[0]) == 12: is_start = True
            elif parts[0] in ["METAR", "TAF", "SPECI"]: is_start = True
            elif len(parts) > 1 and len(parts[0]) == 4 and parts[0].isalpha() and parts[1].endswith('Z'): is_start = True
            
            if parts[0] in ["BECMG", "TEMPO", "PROB30", "PROB40", "RMK"] or parts[0].startswith("FM"):
                is_start = False
        
        if is_start:
            if current_record: data.append(current_record)
            
            ts_raw = parts[0]
            dt_str, turu, content = "---", "METAR", line
            dt_sort = datetime.min
            
            if ts_raw.isdigit() and len(ts_raw) == 12:
                try:
                    dt = datetime.strptime(ts_raw, "%Y%m%d%H%M")
                    dt_str = dt.strftime("%d.%m.%Y %H:%M")
                    dt_sort = dt
                except: pass
                
                if len(parts) > 1:
                    p1 = parts[1]
                    if p1 in ["METAR", "TAF", "SPECI"]:
                        turu = p1
                        content = " ".join(parts[2:])
                    elif p1 == "AAXX":
                        turu = "SİNOPTİK"
                        content = " ".join(parts[1:])
                    else:
                        if "METAR" in line: turu = "METAR"
                        elif "TAF" in line: turu = "TAF"
                        content = " ".join(parts[1:])
            
            elif parts[0] in ["METAR", "TAF", "SPECI"]:
                turu = parts[0]
                content = " ".join(parts[1:])
                m = re.search(r'\b(\d{2})(\d{2})(\d{2})Z\b', content)
                if m:
                    try:
                        now = datetime.now(timezone.utc).replace(tzinfo=None)
                        dt_est = now.replace(day=int(m.group(1)), hour=int(m.group(2)), minute=int(m.group(3)))
                        if dt_est > now + timedelta(days=1): dt_est -= timedelta(days=28)
                        dt_sort = dt_est
                        dt_str = dt_sort.strftime("%d.%m.%Y %H:%M")
                    except: pass

            current_record = {"date": dt_str, "Türü": turu, "İstasyon": station_code, "Bülten": content, "_dt": dt_sort}
        else:
            if current_record: current_record["Bülten"] += " " + line

    if current_record: data.append(current_record)
    
    df = pd.DataFrame(data)
    if not df.empty:
        df = df.drop_duplicates(subset=['Türü', 'Bülten'])
        df = df.sort_values(by="_dt", ascending=False)
    return df

def analyze_dataframe(df):
    """DataFrame üzerindeki METAR ve TAF'ları analiz eder."""
    df["_uyum"] = ""
    df["_detay"] = ""
    df["_ref_taf"] = ""
    
    tafs = df[df['Türü'] == 'TAF'].sort_values(by='_dt')
    if tafs.empty: return df

    last_taf_text = None
    consecutive_counts = {"Rüzgar": 0, "Görüş": 0, "ceil": 0}

    # Kronolojik sıra (Eskiden yeniye)
    for idx, row in df.sort_values(by='_dt', ascending=True).iterrows():
        if row['Türü'] in ['METAR', 'SPECI']:
            metar_dt = row['_dt']
            relevant_tafs = tafs[tafs['_dt'] <= metar_dt]
            
            if not relevant_tafs.empty:
                target_row = relevant_tafs.iloc[-1]
                last_taf = target_row['Bülten']
                taf_dt = target_row['_dt']

                if last_taf != last_taf_text:
                    consecutive_counts = {k: 0 for k in consecutive_counts}
                    last_taf_text = last_taf

                if (metar_dt - taf_dt) > timedelta(hours=3): continue

                df.at[idx, "_ref_taf"] = last_taf
                
                # TAF Zamanı
                regex_period = r'(?:0[1-9]|[12]\d|3[01])(?:[01]\d|2[0-4])/(?:0[1-9]|[12]\d|3[01])(?:[01]\d|2[0-4])'
                t_valid = re.search(r'\b' + regex_period + r'\b', last_taf)
                taf_zaman = t_valid.group(0) if t_valid else "0000/0000"

                # Trend
                trend_part = ""
                tr_m = re.search(r'\b(BECMG|TEMPO|NOSIG)\b', row['Bülten'])
                if tr_m: trend_part = row['Bülten'][tr_m.start():]

                skor, status_code, reasons = robot.analiz_et(last_taf, row['Bülten'], trend_part, taf_zaman)

                # Ardışık Hata Kontrolü
                current_cats = set()
                for r in reasons:
                    if "Rüzgar" in r: current_cats.add("Rüzgar")
                    if "Görüş" in r: current_cats.add("Görüş")
                    if "ceil" in r or "Dikey" in r or "Bulut" in r: current_cats.add("ceil")
                
                amd_msgs = []
                for cat in consecutive_counts:
                    if cat in current_cats:
                        consecutive_counts[cat] += 1
                        if consecutive_counts[cat] >= 3:
                            amd_msgs.append(f"{cat} ({consecutive_counts[cat]}. kez)")
                    else:
                        consecutive_counts[cat] = 0

                icon = ""
                if "UYUMSUZ" in status_code: icon = "❌ UYUMSUZ"
                elif "DİKKAT" in status_code: icon = "⚠️ DİKKAT"
                elif "UYUMLU" in status_code: icon = "✅ UYUMLU"
                
                df.at[idx, "_uyum"] = icon
                
                detay_str = ""
                if "UYUMSUZ" in status_code:
                    detay_str = "1- UYUMSUZLUK NEDENİ:\n" + "\n".join([f"• {r}" for r in reasons])
                    detay_str += "\n\n2- TREND KONTROLÜ:\n• Trend ile de uyum sağlanamadı."
                    detay_str += "\n\n3- SONUÇ:\n• ❌ UYUMSUZ"
                elif "DİKKAT" in status_code:
                    detay_str = "1- UYUMSUZLUK NEDENİ (Ana METAR):\n" + "\n".join([f"• {r}" for r in reasons])
                    detay_str += "\n\n2- TREND KONTROLÜ:\n• ✅ METAR Trendi TAF limitlerine giriyor."
                    detay_str += "\n\n3- SONUÇ:\n• ⚠️ DİKKAT (Trend ile uyumlu)"
                elif "UYUMLU" in status_code:
                    detay_str = "✅ UYUMLU"

                if amd_msgs:
                    detay_str += f"\n\n👉 KRİTİK TAVSİYE:\n• TAF AMD YAYINLANMALI!\n  Aynı sapma 3+ kez tekrarlandı: {', '.join(amd_msgs)}"
                
                df.at[idx, "_detay"] = detay_str

    return df

# --- ARAYÜZ ---
st.subheader("Veri Çekme ve Filtreleme")

col1, col2, col3, col4, col5, col6 = st.columns(6)

with col1:
    station = st.text_input("ICAO Kodu", "LTAN")
with col2:
    wmo = st.text_input("WMO Kodu", "17244")

today = datetime.now()
with col3:
    start_date = st.date_input("Başlangıç", today - timedelta(days=1))
with col4:
    end_date = st.date_input("Bitiş", today)
with col5:
    filter_opts = st.multiselect("Filtrele", ["❌ UYUMSUZ", "⚠️ DİKKAT", "✅ UYUMLU"])

# Session State (Veri Kalıcılığı)
if "analiz_sonucu" not in st.session_state:
    st.session_state.analiz_sonucu = None

with col6:
    st.write("")
    st.write("")
    run_btn = st.button("VERİ ÇEK & ANALİZ ET", type="primary")

if run_btn:
    with st.spinner('Veriler Ogimet üzerinden çekiliyor...'):
        s_dt = datetime.combine(start_date, datetime.min.time())
        e_dt = datetime.combine(end_date, datetime.max.time())
        
        try:
            lines = RASATLAR.fetch(s_dt, e_dt, station=station, wmo_id=wmo)
            if not lines:
                st.error("Veri bulunamadı.")
                st.session_state.analiz_sonucu = None
            else:
                df = process_data(lines, station, wmo)
                df = analyze_dataframe(df)
                st.session_state.analiz_sonucu = df
                            
        except Exception as e:
            st.error(f"Bir hata oluştu: {e}")
            st.session_state.analiz_sonucu = None

if st.session_state.analiz_sonucu is not None:
    df = st.session_state.analiz_sonucu.copy()
    
    # Filtreleme
    if filter_opts:
        df = df[df["_uyum"].isin(filter_opts)]
    
    st.success(f"Toplam {len(df)} kayıt listeleniyor.")
    
    # --- PASTA GRAFİĞİ ---
    if not df.empty:
        st.subheader("Analiz Dağılımı")
        uyum_counts = df["_uyum"].value_counts().reset_index()
        uyum_counts.columns = ["Durum", "Adet"]
        
        color_map = {
            "✅ UYUMLU": "#66BB6A",
            "⚠️ DİKKAT": "#FFEE58",
            "❌ UYUMSUZ": "#EF5350"
        }
        
        col_chart, col_stats = st.columns([2, 1])
        with col_chart:
            fig = px.pie(uyum_counts, values='Adet', names='Durum', 
                         color='Durum', color_discrete_map=color_map,
                         hole=0.4)
            fig.update_layout(height=350, margin=dict(t=10, b=10, l=10, r=10))
            st.plotly_chart(fig, use_container_width=True)
        with col_stats:
            st.dataframe(uyum_counts, hide_index=True, use_container_width=True)
    # ---------------------
    
    # --- EXCEL İNDİRME BUTONU ---
    buffer = io.BytesIO()
    with pd.ExcelWriter(buffer, engine='openpyxl') as writer:
        df.to_excel(writer, index=False, sheet_name='Analiz Raporu')
        
        # Excel Formatlama (Wrap Text ve Sütun Genişliği)
        try:
            from openpyxl.styles import Alignment
            from openpyxl.utils import get_column_letter
            ws = writer.sheets['Analiz Raporu']
            
            for i, col in enumerate(ws.columns, 1):
                col_letter = get_column_letter(i)
                header = str(col[0].value)
                if "Bülten" in header or "Detay" in header:
                    ws.column_dimensions[col_letter].width = 60
                else:
                    ws.column_dimensions[col_letter].width = 15
            
            for row in ws.iter_rows():
                for cell in row:
                    cell.alignment = Alignment(wrap_text=True, vertical='top')
        except: pass
    
    st.download_button(
        label="📥 Excel Olarak İndir",
        data=buffer.getvalue(),
        file_name=f"ogimet_analiz_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    # ---------------------------
    
    # Tablo Gösterimi
    display_df = df[["date", "Türü", "_uyum", "Bülten"]].rename(columns={
        "date": "Tarih",
        "Türü": "Tip",
        "_uyum": "Trend Uyum"
    })
    
    def highlight_rows(row):
        val = str(row["Trend Uyum"])
        if "UYUMSUZ" in val:
            return ['background-color: #ffcdd2; color: black'] * len(row)
        elif "DİKKAT" in val:
            return ['background-color: #fff9c4; color: black'] * len(row)
        elif "UYUMLU" in val:
            return ['background-color: #c8e6c9; color: black'] * len(row)
        return [''] * len(row)

    st.dataframe(
        display_df.style.apply(highlight_rows, axis=1),
        use_container_width=True,
        hide_index=True,
        column_config={"Bülten": st.column_config.TextColumn("Bülten", width="large")}
    )
    
    # Detaylar
    with st.expander("Detaylı Analiz Raporu"):
        for _, row in df.iterrows():
            if row["_detay"]:
                st.markdown(f"**{row['date']} - {row['Bülten']}**")
                st.info(row["_detay"])
                st.divider()